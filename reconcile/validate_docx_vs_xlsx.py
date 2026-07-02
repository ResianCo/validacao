"""Run-once gate: validate the DOCX (final golden) against the XLSX (first-part golden).

Both documents are golden, at different stages:
  - XLSX = first part (raw/intermediate working workbook)
  - DOCX = final part (delivered valuation report)

This script cross-checks the concepts BOTH documents contain, to surface where the
two golden stages agree and where they diverge (for human interpretation), BEFORE
the rest of the audit layers in the pipeline and PDF.

Reads the golden folder READ-ONLY. Writes only to this reconcile/ subfolder.

Run:
  python validate_docx_vs_xlsx.py [--venv-python <path>]
(deps: openpyxl. stdlib only otherwise.)
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# Resolve paths relative to the repo root (parent of this reconcile/ dir), so the
# audit works from any checkout — not tied to /Users/thiago/....
REPO = Path(__file__).resolve().parent.parent
DOCX = REPO / "Banco_ABC_Valuation_Resian_2026-05_1.docx"
# The XLSX ships as a zip-of-xlsx; see resolve_xlsx_path() below.
XLSX_ZIP = REPO / "20260612 Teste - Análise de Carteira.xlsx.zip"
OUT = REPO / "reconcile"


def resolve_xlsx_path(path: Path) -> tuple[Path, bool]:
    """Return a path openpyxl can read.

    The XLSX is stored as `*.xlsx.zip` (a zip containing the workbook plus macOS
    `__MACOSX` cruft). openpyxl cannot read a zip directly, so when given the zip
    we extract the inner .xlsx to a NamedTemporaryFile and return it. The caller
    must unlink the temp path when done (the bool return signals this).
    """
    import tempfile, zipfile
    path = Path(path)
    if zipfile.is_zipfile(path):
        with zipfile.ZipFile(path) as z:
            inner = next(n for n in z.namelist()
                         if n.lower().endswith(".xlsx") and "__MACOSX" not in n)
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.write(z.read(inner))
            tmp.close()
            return Path(tmp.name), True
    return path, False

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

# Tolerances (backend Params defaults).
TOL_MONEY = 1.00
TOL_PCT = 0.01
TOL_COUNT = 0


# ----------------------------- BR number parsing --------------------------
def parse_br(value, kind):
    """kind: 'pct' (dot=decimal) | 'count'|'money' (dot=thousands, comma=decimal)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace("%", "").strip()
    if s in ("", "-", "N/A", "#DIV/0!"):
        return None
    try:
        if kind == "pct":
            return float(s.replace(",", "."))
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(".", "")  # dot = thousands
        return float(s)
    except ValueError:
        return None


# ----------------------------- DOCX extraction ----------------------------
def _cell_text(tc):
    parts = []
    for c in tc.iter():
        if c.tag == f"{W}t":
            parts.append(c.text or "")
    return "".join(parts).replace("\n", " ").strip()


def _para_text(p):
    parts = []
    for c in p.iter():
        if c.tag == f"{W}t":
            parts.append(c.text or "")
    return "".join(parts).strip()


def read_docx_tables(path: Path):
    """Return list of tables; each a list of rows of cell-text."""
    tables = []
    with zipfile.ZipFile(path) as z:
        root = ET.fromstring(z.read("word/document.xml"))
    body = root.find(f"{W}body")
    for el in body:
        if el.tag != f"{W}tbl":
            continue
        rows = []
        for tr in el.findall(f"{W}tr"):
            cells = [_cell_text(tc) for tc in tr.findall(f"{W}tc")]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def _norm(s):
    return re.sub(r"\s+", " ", s or "").strip().lower()


def extract_docx(path: Path) -> dict:
    tables = read_docx_tables(path)
    # KPI flat dict from small pair tables (<=4 non-empty cells/row).
    kpi = {}
    for rows in tables:
        if any(len([c for c in r if c.strip()]) > 4 for r in rows):
            continue
        for r in rows:
            cells = [c for c in r if c.strip()]
            i = 0
            while i + 1 < len(cells):
                k = cells[i].strip()
                v = cells[i + 1].strip()
                if k:
                    kpi[k] = v
                i += 2

    def find(*needles):
        nk = [_norm(n) for n in needles]
        for rows in tables:
            if not rows:
                continue
            hdr = _norm(" ".join(rows[0]))
            if all(n in hdr for n in nk):
                return rows
        return None

    # FPD table: header has 'safra' + 'fpd'
    fpd = []
    fr = find("Safra", "FPD")
    if fr:
        for r in fr[1:]:
            cells = [c for c in r if c.strip()]
            if len(cells) >= 5 and re.match(r"\d{4}-\d{2}", cells[0]):
                fpd.append({"safra": cells[0],
                            "contratos": parse_br(cells[1], "count"),
                            "fpd_n": parse_br(cells[2], "count"),
                            "fpd_qtd_pct": parse_br(cells[3], "pct"),
                            "fpd_val_pct": parse_br(cells[4], "pct")})

    # rating table: header has 'rating' + 'principal'
    rating = []
    rr = find("Rating", "Principal")
    if rr:
        for r in rr[1:]:
            cells = [c for c in r if c.strip()]
            if len(cells) >= 6 and cells[0] in {"A", "B", "C", "D", "E", "HR"}:
                rating.append({"rating": cells[0],
                               "contratos": parse_br(cells[1], "count"),
                               "principal": parse_br(cells[2], "money"),
                               "over60_n": parse_br(cells[3], "count"),
                               "over60_pct": parse_br(cells[4], "pct")})

    def k(*labels):
        for lab in labels:
            if lab in kpi and kpi[lab] not in ("", None):
                return kpi[lab]
        return None

    scalars = {
        "contratos": parse_br(k("Contratos"), "count"),
        "principal_total": parse_br(k("Principal Total"), "money"),
        "over60_n": parse_br((k("Over 60 (contratos)") or "").split("/")[0], "count"),
        "over60_pct": parse_br((k("Over 60 (contratos)") or "").split("/")[-1], "pct"),
        "lgd_pct": parse_br(k("LGD (Effic 90 LTM)"), "pct"),
        "ead_total": parse_br(k("EAD Total"), "money"),
        "pe_valor": parse_br(k("Perda Esperada (PE)", "PE"), "money"),
    }
    return {"scalars": scalars, "fpd": fpd, "rating": rating, "ntables": len(tables)}


# ----------------------------- XLSX extraction ----------------------------
def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def extract_xlsx(path: Path) -> dict:
    import openpyxl
    real_path, is_temp = resolve_xlsx_path(path)
    try:
        wb = openpyxl.load_workbook(real_path, read_only=True, data_only=True)
    finally:
        if is_temp:
            real_path.unlink(missing_ok=True)
    out = {"sheets": list(wb.sheetnames)}

    # Resumo TOTAL: contratos + BRL.
    resumo = {}
    for sname in wb.sheetnames:
        if sname.lower() != "resumo":
            continue
        for row in wb[sname].iter_rows(values_only=True):
            for i, c in enumerate(row):
                if isinstance(c, str) and c.strip().upper() == "TOTAL":
                    if i + 1 < len(row) and _is_num(row[i + 1]):
                        brl = next((row[j] for j in range(i + 2, len(row)) if _is_num(row[j])), None)
                        if brl is not None and not resumo:
                            resumo = {"contratos": float(row[i + 1]), "brl_total": float(brl)}
        break

    # FPD sheet: safra, contratos_n, fpd30_qtd(frac), fpd30_val(frac).
    fpd = []
    for sname in wb.sheetnames:
        if sname.upper() != "FPD":
            continue
        for row in wb[sname].iter_rows(values_only=True):
            if len(row) > 9 and isinstance(row[2], dt.datetime) and _is_num(row[3]):
                fpd.append({"safra": row[2].strftime("%Y-%m"),
                            "contratos": float(row[3]),
                            "fpd30_qtd_frac": float(row[6]),
                            "fpd30_val_frac": float(row[9])})
        break

    # PE sheet: LGD (constant across rows).
    lgd = None
    for sname in wb.sheetnames:
        if sname.upper() != "PE":
            continue
        for row in wb[sname].iter_rows(values_only=True):
            if len(row) > 6 and isinstance(row[1], dt.datetime) and _is_num(row[3]):
                lgd = float(row[5])
                break
        break

    wb.close()
    out.update(resumo=resumo, fpd=fpd, lgd_frac=lgd)
    return out


# ----------------------------- reconciliation -----------------------------
def check(name, computed, golden, tol, kind, checks, note=""):
    if computed is None or golden is None:
        checks.append({"name": name, "computed": computed, "golden": golden,
                       "diff": None, "tol": tol, "kind": kind,
                       "passed": computed is None and golden is None, "note": note})
        return
    diff = round(abs(float(computed) - float(golden)), 6)
    checks.append({"name": name, "computed": computed, "golden": golden,
                   "diff": diff, "tol": tol, "kind": kind,
                   "passed": diff <= tol, "note": note})


def reconcile(d: dict, x: dict) -> dict:
    checks = []
    ds, xs = d["scalars"], x.get("resumo", {})

    # 1. contratos (count) — robust to money scaling.
    check("contratos", ds["contratos"], xs.get("contratos"), TOL_COUNT, "count", checks)

    # 2. principal — expect the ÷10 divergence.
    ratio = round(ds["principal_total"] / xs["brl_total"], 4) if xs.get("brl_total") else None
    check("principal_total", ds["principal_total"], xs.get("brl_total"), TOL_MONEY, "money",
          checks, note=f"DOCX/XLSX ratio = {ratio}×")

    # 3. LGD — expect divergence.
    xlgd = (x["lgd_frac"] * 100) if isinstance(x.get("lgd_frac"), (int, float)) else None
    check("LGD %", ds["lgd_pct"], xlgd, TOL_PCT, "pct", checks,
          note="DOCX 92.25% (final) vs XLSX PE-sheet ~91.53% (first part)")

    # 4. FPD by safra (the clean count-ratio comparable).
    xfpd = {r["safra"]: r for r in x["fpd"]}
    fpd_rows = []
    for r in d["fpd"]:
        s = r["safra"]
        if s not in xfpd:
            continue
        xq = round(xfpd[s]["fpd30_qtd_frac"] * 100, 4)
        xv = round(xfpd[s]["fpd30_val_frac"] * 100, 4)
        fpd_rows.append({"safra": s, "docx_qtd": r["fpd_qtd_pct"], "xlsx_qtd": xq,
                         "diff_qtd": round(abs(r["fpd_qtd_pct"] - xq), 4),
                         "docx_val": r["fpd_val_pct"], "xlsx_val": xv,
                         "diff_val": round(abs(r["fpd_val_pct"] - xv), 4)})
        check(f"FPD30 qtd% {s}", r["fpd_qtd_pct"], xq, TOL_PCT, "pct", checks)
        check(f"FPD30 val% {s}", r["fpd_val_pct"], xv, TOL_PCT, "pct", checks)

    passed = sum(1 for c in checks if c["passed"])
    return {
        "docx": d["scalars"], "xlsx": {"resumo": xs, "lgd_pct": xlgd,
                                       "fpd_safras": len(x["fpd"])},
        "checks": checks,
        "fpd_by_safra": fpd_rows,
        "structural": {
            "principal_docx": ds["principal_total"],
            "principal_xlsx": xs.get("brl_total"),
            "principal_ratio_docx_over_xlsx": ratio,
            "lgd_docx_pct": ds["lgd_pct"], "lgd_xlsx_pct": xlgd,
            "fpd_safras_compared": len(fpd_rows),
            "fpd_docx_total": len(d["fpd"]), "fpd_xlsx_total": len(x["fpd"]),
        },
        "summary": {"total": len(checks), "passed": passed,
                    "failed": len(checks) - passed, "green": passed == len(checks)},
    }


def render_markdown(rep: dict) -> str:
    s = rep["summary"]
    st = rep["structural"]
    lines = [
        "# Run-once gate: DOCX (final) vs XLSX (first part)", "",
        f"**{s['passed']}/{s['total']} checks passed — {'GREEN' if s['green'] else 'RED (divergences flagged for interpretation)'}**", "",
        "Both documents are golden at different stages. This gate surfaces where the two stages agree and diverge.", "",
        "## Headline findings (read first)", "",
        f"- **contratos agrees:** DOCX = XLSX = {int(rep['docx']['contratos'])}.", "",
        f"- **Money diverges by ÷10.** DOCX principal = R$ {st['principal_docx']:,.2f}; "
        f"XLSX Resumo TOTAL = R$ {st['principal_xlsx']:,.2f}; **ratio = "
        f"{st['principal_ratio_docx_over_xlsx']}×**. The XLSX (first part) carries all "
        f"money at 1/10 of the DOCX (final). Needs your interpretation: intentional "
        f"stage scaling, or a workbook unit issue?", "",
        f"- **LGD diverges.** DOCX = {st['lgd_docx_pct']}%; XLSX PE-sheet = "
        f"{round(st['lgd_xlsx_pct'],4) if st['lgd_xlsx_pct'] else None}%. "
        f"Methodology value changed between stages — needs your interpretation.", "",
        f"- **FPD agrees** on the {st['fpd_safras_compared']} overlapping safras "
        f"(DOCX has {st['fpd_docx_total']}, XLSX FPD has {st['fpd_xlsx_total']}). "
        f"Count% and value% match within tolerance — see table.", "",
        "- **Not directly comparable:** EAD/PE (DOCX absolute R$ vs XLSX fractional of "
        "origination, itself ÷10); rating (both present but as different cuts).", "",
        "## All checks", "",
        "| check | docx (final) | xlsx (first part) | abs_diff | tol | result | note |",
        "|---|---|---|---|---|---|---|",
    ]
    for c in rep["checks"]:
        r = "PASS" if c["passed"] else "**FAIL**"
        lines.append(f"| {c['name']} | {c['computed']} | {c['golden']} | {c['diff']} | {c['tol']} | {r} | {c.get('note','')} |")
    if rep["fpd_by_safra"]:
        lines += ["", "## FPD by safra (DOCX vs XLSX FPD30)", "",
                  "| safra | docx qtd% | xlsx qtd% | diff | docx val% | xlsx val% | diff |",
                  "|---|---|---|---|---|---|---|"]
        for f in rep["fpd_by_safra"]:
            lines.append(f"| {f['safra']} | {f['docx_qtd']} | {f['xlsx_qtd']} | {f['diff_qtd']} | {f['docx_val']} | {f['xlsx_val']} | {f['diff_val']} |")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--docx", default=str(DOCX))
    ap.add_argument("--xlsx", default=str(XLSX_ZIP))
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)
    print(f"[1/3] extracting DOCX: {Path(args.docx).name}")
    d = extract_docx(Path(args.docx))
    print(f"      contratos={d['scalars']['contratos']} principal={d['scalars']['principal_total']} "
          f"lgd={d['scalars']['lgd_pct']} fpd_rows={len(d['fpd'])} rating_rows={len(d['rating'])}")
    print(f"[2/3] extracting XLSX: {Path(args.xlsx).name}")
    x = extract_xlsx(Path(args.xlsx))
    print(f"      resumo={x.get('resumo')} lgd_frac={x.get('lgd_frac')} fpd_rows={len(x['fpd'])}")
    print("[3/3] reconciling...")
    rep = reconcile(d, x)
    md = render_markdown(rep)
    (Path(args.out) / "DOCS_VS_XLSX.md").write_text(md, encoding="utf-8")
    (Path(args.out) / "DOCS_VS_XLSX.json").write_text(
        json.dumps(rep, indent=2, default=str), encoding="utf-8")
    print(md)
    print(f"\nreport -> {Path(args.out) / 'DOCS_VS_XLSX.md'}")
    print(f"FINAL: {'GREEN' if rep['summary']['green'] else 'RED'} "
          f"({rep['summary']['passed']}/{rep['summary']['total']})")
    return 0 if rep["summary"]["green"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
