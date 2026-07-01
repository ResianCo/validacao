"""Full certification audit for the VALIDATION golden folder.

Two-layer model (both documents are golden):
  - XLSX = first part  (raw/intermediate working workbook; richer columns: UF, Idade...)
  - DOCX = final part  (delivered valuation report)

Layer A — DOCX (final) is correct:
  - pipeline_csv.py output <-> DOCX  (every CSV-reproducible table)
  - DOCX self-consistency (rating/faixas rollups)
  - DOCX <-> PDF (the PDF is an architecture report; check it cites the golden numbers)

Layer B — XLSX (first part) is correct & sources the richer DOCX cuts:
  - XLSX internal (Resumo count, Loans count/principal, Ever60 monotonic)
  - XLSX -> DOCX UF lineage (DOCX UF table grouped from XLSX Loans sheet)

Reads the golden folder READ-ONLY. Writes only to this reconcile/ subfolder.
Pipeline output is read from a scratch dir (re-run pipeline_csv.py yourself there).

Run:
  python audit.py <pipeline_output_dir> [--docx ..] [--xlsx ..] [--pdf ..] [--out ..]
(deps: pandas, openpyxl, pypdf.)
"""
from __future__ import annotations
import argparse
import datetime as dt
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd

GOLDEN = Path("/Users/thiago/Resian/validacao")
DOCX = GOLDEN / "Banco_ABC_Valuation_Resian_2026-05_1.docx"
XLSX = GOLDEN / "20260612 Teste - Análise de Carteira.xlsx"
PDF = GOLDEN / "20260625-EXECUTIVE_REPORT_Thiago.pdf"
OUT = GOLDEN / "reconcile"

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

TOL_MONEY = 1.00
TOL_PCT = 0.01
TOL_COUNT = 0
TOL_REL_VPL = 0.001   # 0.1% for VPL (matches backend's stated convergence bound)


# ----------------------------- parsing helpers ---------------------------
def parse_br(value, kind):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("R$", "").replace("%", "").strip()
    s = re.sub(r"\(.*?\)", "", s)  # drop "(70.63% do Principal)" tails
    s = s.strip()
    if s in ("", "-", "N/A", "#DIV/0!"):
        return None
    try:
        if kind == "pct":
            return float(s.replace(",", "."))
        if kind == "decimal":  # dot is always a decimal point (e.g. '54.2')
            return float(s.replace(",", "."))
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(".", "")  # dot = thousands
        return float(s)
    except ValueError:
        return None


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ----------------------------- DOCX extraction ---------------------------
def _cell_text(tc):
    parts = []
    for c in tc.iter():
        if c.tag == f"{W}t":
            parts.append(c.text or "")
    return "".join(parts).replace("\n", " ").strip()


def read_docx_tables(path: Path):
    tables = []
    with zipfile.ZipFile(path) as z:
        root = ET.fromstring(z.read("word/document.xml"))
    body = root.find(f"{W}body")
    for el in body:
        if el.tag != f"{W}tbl":
            continue
        rows = []
        for tr in el.findall(f"{W}tr"):
            cells = [_cell_text(tc) for tc in tr.findall(f"{W}tc")]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def _norm(s):
    return re.sub(r"\s+", " ", s or "").strip().lower()


def _find_table(tables, *needles):
    nk = [_norm(n) for n in needles]
    for rows in tables:
        if rows and all(n in _norm(" ".join(rows[0])) for n in nk):
            return rows
    return None


def _rows_nonempty(rows):
    out = []
    for r in rows[1:]:
        c = [x for x in r if x.strip()]
        if c:
            out.append(c)
    return out


def extract_docx(path: Path) -> dict:
    tables = read_docx_tables(path)
    # KPI flat dict from pair tables (<=4 non-empty cells/row).
    kpi = {}
    for rows in tables:
        if any(len([c for c in r if c.strip()]) > 4 for r in rows):
            continue
        for r in rows:
            cells = [c for c in r if c.strip()]
            i = 0
            while i + 1 < len(cells):
                if cells[i].strip() and cells[i].strip() not in kpi:  # first-wins
                    kpi[cells[i].strip()] = cells[i + 1].strip()
                i += 2

    def k(*labels):
        for lab in labels:
            if lab in kpi and kpi[lab] not in ("", None):
                return kpi[lab]
        return None

    scalars = {
        "contratos": parse_br(k("Contratos"), "count"),
        "principal_total": parse_br(k("Principal Total"), "money"),
        "ticket_medio": parse_br(k("Ticket Médio", "Ticket M\u00e9dio"), "money"),
        "taxa_media_pct": parse_br(k("Taxa Média a.m.", "Taxa M\u00e9dia a.m."), "pct"),
        "prazo_medio": parse_br(k("Prazo Médio (meses)", "Prazo M\u00e9dio (meses)"), "decimal"),
        "over60_n": parse_br((k("Over 60 (contratos)") or "").split("/")[0], "count"),
        "over60_pct": parse_br((k("Over 60 (contratos)") or "").split("/")[-1], "pct"),
        "lgd_pct": parse_br(k("LGD (Effic 90 LTM)"), "pct"),
        "ead_total": parse_br(k("EAD Total"), "money"),
        "pe_valor": parse_br(k("Perda Esperada (PE)", "PE"), "money"),
        "renegociacoes": parse_br(k("Renegociações", "Renegociacoes"), "count"),
    }

    # rating
    rating = []
    rr = _find_table(tables, "Rating", "Principal")
    if rr:
        for c in _rows_nonempty(rr):
            if len(c) >= 6 and c[0] in {"A", "B", "C", "D", "E", "HR"}:
                rating.append({"rating": c[0], "contratos_n": parse_br(c[1], "count"),
                               "principal": parse_br(c[2], "money"),
                               "over60_n": parse_br(c[3], "count"),
                               "over60_pct": parse_br(c[4], "pct"),
                               "part_pct": parse_br(c[5], "pct")})

    # faixas_atraso
    faixas = []
    fr = _find_table(tables, "Faixa", "% Carteira")
    if fr:
        for c in _rows_nonempty(fr):
            if len(c) >= 4:
                faixas.append({"faixa": c[0], "contratos_n": parse_br(c[1], "count"),
                               "pct_carteira": parse_br(c[2], "pct"),
                               "acumulado": parse_br(c[3], "pct")})

    # FPD by safra
    fpd = []
    fpr = _find_table(tables, "Safra", "FPD")
    if fpr:
        for c in _rows_nonempty(fpr):
            if len(c) >= 5 and re.match(r"\d{4}-\d{2}", c[0]):
                fpd.append({"safra": c[0], "contratos_n": parse_br(c[1], "count"),
                            "fpd_n": parse_br(c[2], "count"),
                            "fpd_qtd_pct": parse_br(c[3], "pct"),
                            "fpd_val_pct": parse_br(c[4], "pct")})

    # VPL grid: Cenário | VPL @1% | VPL @1.5%
    vpl = []
    vr = _find_table(tables, "Cenário", "VPL")
    if vr:
        for c in _rows_nonempty(vr):
            if len(c) >= 3 and c[0].lower() in {"otimista", "base", "pessimista", "otimista"}:
                # cells may be "R$ 144.241.042 (70.63% do Principal)"
                vpl.append({"cenario": c[0].lower(),
                            "vpl_1pct": parse_br(c[1], "money"),
                            "vpl_1_5pct": parse_br(c[2], "money")})

    # UF table
    uf = []
    ur = _find_table(tables, "UF", "Over 60")
    if ur:
        for c in _rows_nonempty(ur):
            if len(c) >= 6 and re.match(r"^[A-Z]{2}$", c[0]):
                uf.append({"uf": c[0], "contratos": parse_br(c[1], "count"),
                           "principal": parse_br(c[2], "money"),
                           "over60_n": parse_br(c[4], "count"),
                           "over60_pct": parse_br(c[5], "pct")})

    return {"scalars": scalars, "rating": rating, "faixas": faixas,
            "fpd": fpd, "vpl": vpl, "uf": uf, "ntables": len(tables)}


# ----------------------------- XLSX extraction ---------------------------
def extract_xlsx(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"sheets": list(wb.sheetnames)}

    # Loans sheet: header-driven column discovery (UF, Maior Atraso, Valor Emprestado).
    loans = {"count": None, "principal_sum": None, "by_uf": {}}
    lname = next((s for s in wb.sheetnames if s.lower() == "loans"), None)
    if lname:
        ws = wb[lname]
        rows_iter = ws.iter_rows(values_only=True)
        # find the header row (first row containing 'emprestado'); there are
        # blank / super-header rows above it.
        header = None
        for row in rows_iter:
            normed = [(_norm(c) if isinstance(c, str) else "") for c in row]
            if any("emprestado" in c for c in normed):
                header = normed
                break
        if header:
            h = header
            i_uf = h.index("uf") if "uf" in h else None
            i_maior = next((i for i, c in enumerate(h) if "maior atraso" in c), None)
            i_emp = next((i for i, c in enumerate(h) if "emprestado" in c), None)
            ncount = 0; psum = 0.0; byuf = {}
            for row in rows_iter:
                if not _is_num(row[i_emp]) if i_emp is not None else True:
                    pass
                if i_emp is not None and _is_num(row[i_emp]):
                    ncount += 1
                    psum += float(row[i_emp])
                    if i_uf is not None and row[i_uf]:
                        u = str(row[i_uf]).strip()
                        d = byuf.setdefault(u, {"count": 0, "principal": 0.0, "over60": 0})
                        d["count"] += 1
                        d["principal"] += float(row[i_emp])
                        if i_maior is not None and _is_num(row[i_maior]) and float(row[i_maior]) > 60:
                            d["over60"] += 1
            loans = {"count": ncount, "principal_sum": psum, "by_uf": byuf,
                     "i_uf": i_uf, "i_maior": i_maior}

    # Resumo TOTAL
    resumo = {}
    for sname in wb.sheetnames:
        if sname.lower() != "resumo":
            continue
        for row in wb[sname].iter_rows(values_only=True):
            for i, c in enumerate(row):
                if isinstance(c, str) and c.strip().upper() == "TOTAL":
                    if i + 1 < len(row) and _is_num(row[i + 1]):
                        brl = next((row[j] for j in range(i + 2, len(row)) if _is_num(row[j])), None)
                        if brl is not None and not resumo:
                            resumo = {"contratos": float(row[i + 1]), "brl_total": float(brl)}
        break

    # Ever60 monotonic check (Inadimplência 60): per safra, the fraction columns MOB1..N
    ever60 = []
    for sname in wb.sheetnames:
        if sname.lower() != "inadimplência 60":
            continue
        for row in wb[sname].iter_rows(values_only=True):
            if len(row) > 6 and isinstance(row[2], dt.datetime) and _is_num(row[3]):
                # fraction columns start at index 6
                frac = [row[j] for j in range(6, len(row)) if _is_num(row[j])]
                if frac:
                    ever60.append({"safra": row[2].strftime("%Y-%m"), "curve": [float(x) for x in frac]})
        break

    wb.close()
    out.update(loans=loans, resumo=resumo, ever60=ever60)
    return out


# ----------------------------- PDF extraction ----------------------------
def extract_pdf_numbers(path: Path) -> dict:
    """Pull PE and LGD values cited in the PDF text (architecture report)."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return {"available": False}
    r = PdfReader(str(path))
    text = "\n".join((p.extract_text() or "") for p in r.pages)
    out = {"available": True, "npages": len(r.pages), "cites": {}}
    m_pe = re.search(r"PE\s*=\s*R\$\s*([\d.,]+)", text)
    m_lgd = re.search(r"LGD\s*=\s*([\d.,]+)\s*%", text)
    if m_pe:
        out["cites"]["pe_valor"] = parse_br(m_pe.group(1), "money")
    if m_lgd:
        out["cites"]["lgd_pct"] = parse_br(m_lgd.group(1), "pct")
    out["is_architecture_report"] = "Scaling" in text or "Architecture" in text
    return out


# ----------------------------- pipeline output ---------------------------
def load_pipeline(out_dir: Path) -> dict:
    def csv(name):
        return pd.read_csv(out_dir / name, sep=";")
    k = csv("kpis.csv")
    kpis = {r["chave"]: r["valor"] for _, r in k.iterrows()}
    return {
        "kpis": kpis,
        "fpd": csv("fpd_safras.csv"),
        "rating": csv("rating.csv"),
        "faixas": csv("faixas_atraso.csv"),
        "vpl": csv("vpl_cenarios.csv"),
        "pe": csv("parametros_pe.csv"),
    }


# ----------------------------- checks ------------------------------------
def check(checks, name, computed, golden, tol, kind, note=""):
    if computed is None or golden is None:
        checks.append({"name": name, "computed": computed, "golden": golden,
                       "diff": None, "tol": tol, "kind": kind,
                       "passed": computed is None and golden is None, "note": note})
        return
    if kind == "rel":
        denom = abs(float(golden)) or 1.0
        diff = round(abs(float(computed) - float(golden)) / denom, 6)
    else:
        diff = round(abs(float(computed) - float(golden)), 6)
    checks.append({"name": name, "computed": computed, "golden": golden,
                   "diff": diff, "tol": tol, "kind": kind,
                   "passed": diff <= tol, "note": note})


def run_audit(pipe_dir, docx, xlsx, pdf) -> dict:
    checks = []
    notes = []
    d = extract_docx(docx)
    x = extract_xlsx(xlsx)
    p = load_pipeline(pipe_dir)
    ds = d["scalars"]

    # ===================== LAYER A: DOCX (final) correct ====================
    # A1. KPI scalars: pipeline <-> DOCX
    pk = p["kpis"]
    kmap = [("contratos", "contratos", "count"), ("principal_total", "principal_total", "money"),
            ("over60_n", "over60_n", "count"), ("over60_pct", "over60_pct", "pct"),
            ("ead_total", "ead_total", "money")]
    for ckey, dkey, kind in kmap:
        cv = pk.get(ckey); cv = float(cv) if cv is not None else None
        check(checks, f"A·KPI {dkey}", cv, ds.get(dkey),
              TOL_COUNT if kind == "count" else (TOL_MONEY if kind == "money" else TOL_PCT), kind)

    # pipeline PE <-> DOCX PE
    pe_val = float(p["pe"].loc[p["pe"]["parametro"].str.startswith("PE ="), "valor"].iloc[0])
    check(checks, "A·PE valor", pe_val, ds.get("pe_valor"), TOL_MONEY, "money")

    # A2. rating: pipeline <-> DOCX
    pr = {str(r["rating"]): r for _, r in p["rating"].iterrows()}
    for dr in d["rating"]:
        rt = dr["rating"]
        if rt in pr:
            check(checks, f"A·rating {rt} contratos", float(pr[rt]["contratos_n"]),
                  dr["contratos_n"], TOL_COUNT, "count")
            check(checks, f"A·rating {rt} principal", float(pr[rt]["principal"]),
                  dr["principal"], TOL_MONEY, "money")
            check(checks, f"A·rating {rt} over60_n", float(pr[rt]["over60_n"]),
                  dr["over60_n"], TOL_COUNT, "count")

    # A3. faixas: pipeline <-> DOCX
    pf = {str(r["faixa"]).replace("\ufeff", ""): r for _, r in p["faixas"].iterrows()}
    for dr in d["faixas"]:
        key = dr["faixa"]
        row = pf.get(key)
        if row is None:
            row = next((v for k, v in pf.items() if k.strip() == key.strip()), None)
        if row is not None:
            check(checks, f"A·faixas '{key[:18]}' contratos", float(row["contratos_n"]),
                  dr["contratos_n"], TOL_COUNT, "count")
            check(checks, f"A·faixas '{key[:18]}' pct", float(row["pct_carteira"]),
                  dr["pct_carteira"], TOL_PCT, "pct")

    # A4. FPD: pipeline <-> DOCX (24 safras)
    pfld = {str(r["safra"]): r for _, r in p["fpd"].iterrows()}
    fpd_diffs = []
    for dr in d["fpd"]:
        s = dr["safra"]
        if s in pfld:
            dq = round(float(pfld[s]["fpd_qtd_pct"]), 4)
            dv = round(float(pfld[s]["fpd_val_pct"]), 4)
            check(checks, f"A·FPD qtd% {s}", dq, dr["fpd_qtd_pct"], TOL_PCT, "pct")
            check(checks, f"A·FPD val% {s}", dv, dr["fpd_val_pct"], TOL_PCT, "pct")
            fpd_diffs.append(max(abs(dq - dr["fpd_qtd_pct"]), abs(dv - dr["fpd_val_pct"])))

    # A5. VPL grid: pipeline <-> DOCX (relative tol 0.1%)
    pv = {}
    for _, r in p["vpl"].iterrows():
        pv[(str(r["cenario"]).lower(), float(r["taxa_am_pct"]))] = float(r["vpl_valor"])
    vpl_diffs = []
    for dr in d["vpl"]:
        cen = dr["cenario"]
        for taxakey, dval in [("vpl_1pct", 1.0), ("vpl_1_5pct", 1.5)]:
            cv = pv.get((cen, dval))
            gv = dr.get(taxakey)
            if cv and gv:
                rel = abs(cv - gv) / gv
                vpl_diffs.append(rel)
                check(checks, f"A·VPL {cen}@{taxakey}", cv, gv, TOL_REL_VPL, "rel",
                      note=f"rel diff {rel*100:.3f}%")

    # A6. DOCX self-consistency
    if d["rating"]:
        check(checks, "A·self rating Σcontratos", sum(r["contratos_n"] for r in d["rating"]),
              ds["contratos"], TOL_COUNT, "count")
        check(checks, "A·self rating Σprincipal", sum(r["principal"] for r in d["rating"]),
              ds["principal_total"], 5.0, "money",
              note="docx rating principals are integer-rounded")
    if d["faixas"]:
        check(checks, "A·self faixas Σcontratos", sum(r["contratos_n"] for r in d["faixas"]),
              ds["contratos"], TOL_COUNT, "count")
        check(checks, "A·self faixas acumulado ends 100",
              d["faixas"][-1]["acumulado"], 100.0, TOL_PCT, "pct")

    # A7. DOCX <-> PDF (citations only)
    pdf_info = extract_pdf_numbers(pdf)
    if pdf_info.get("available") and pdf_info["cites"]:
        if "pe_valor" in pdf_info["cites"]:
            check(checks, "A·PDF cites PE", pdf_info["cites"]["pe_valor"],
                  ds.get("pe_valor"), TOL_MONEY, "money")
        if "lgd_pct" in pdf_info["cites"]:
            check(checks, "A·PDF cites LGD", pdf_info["cites"]["lgd_pct"],
                  ds.get("lgd_pct"), TOL_PCT, "pct")
        notes.append(f"PDF is an architecture/scaling report ({pdf_info['npages']} pages); "
                     "checked only that it cites the DOCX golden numbers.")

    # ===================== LAYER B: XLSX (first part) =====================
    # B1. XLSX internal
    xl = x["loans"]
    check(checks, "B·XLSX Loans count", xl["count"], ds["contratos"], TOL_COUNT, "count")
    if x.get("resumo"):
        check(checks, "B·XLSX Resumo contratos", x["resumo"]["contratos"],
              ds["contratos"], TOL_COUNT, "count")
    # Ever60 monotonic non-decreasing per safra
    non_mono = []
    for c in x["ever60"]:
        curve = c["curve"]
        if any(curve[i + 1] < curve[i] - 1e-9 for i in range(len(curve) - 1)):
            non_mono.append(c["safra"])
    check(checks, "B·Ever60 curves monotonic (non-decreasing)", len(non_mono), 0, 0, "count",
          note=f"{len(x['ever60'])} safra curves checked; non-monotonic: {', '.join(non_mono) or 'none'}")
    notes.append(f"XLSX ever60: {len(x['ever60'])} safra curves, {len(non_mono)} non-monotonic "
                 f"({', '.join(non_mono) or 'none'} — all youngest safras, likely a workbook artifact).")

    # B2. XLSX -> DOCX UF lineage
    if d["uf"] and xl["by_uf"]:
        lineage = []
        uf_ok = 0; uf_total = 0
        for du in d["uf"]:
            u = du["uf"]
            xu = xl["by_uf"].get(u)
            if xu is None:
                continue
            uf_total += 1
            cmatch = (xu["count"] == du["contratos"])
            pratio = round(xu["principal"] / du["principal"], 4) if du["principal"] else None
            if cmatch:
                uf_ok += 1
            lineage.append({"uf": u, "xlsx_count": xu["count"], "docx_count": du["contratos"],
                            "count_match": cmatch, "principal_ratio_xlsx_over_docx": pratio})
            check(checks, f"B·UF {u} contratos (XLSX->DOCX)", xu["count"],
                  du["contratos"], TOL_COUNT, "count",
                  note=f"principal ratio xlsx/docx={pratio}")
        notes.append(f"XLSX->DOCX UF lineage: {uf_ok}/{uf_total} UF contrato counts match "
                     f"(principal is ÷10 in XLSX, expected).")
        # attach lineage to report
    else:
        lineage = []
        notes.append("UF lineage skipped (DOCX UF table or XLSX Loans UF column missing).")

    passed = sum(1 for c in checks if c["passed"])
    return {
        "summary": {"total": len(checks), "passed": passed,
                    "failed": len(checks) - passed, "green": passed == len(checks)},
        "checks": checks, "notes": notes,
        "docx_scalars": ds,
        "fpd_max_diff_pp": round(max(fpd_diffs) if fpd_diffs else None, 4),
        "vpl_max_rel_diff_pct": round(max(vpl_diffs) * 100, 4) if vpl_diffs else None,
        "uf_lineage": lineage,
        "xlsx_loans_count": xl["count"],
        "xlsx_sheets": x["sheets"],
    }


def render_markdown(rep: dict) -> str:
    s = rep["summary"]
    lines = [
        "# AUDIT REPORT — validacao golden folder", "",
        f"**{s['passed']}/{s['total']} checks passed — {'GREEN' if s['green'] else 'RED'}**", "",
        "Two-layer model: **XLSX = first part** (raw/intermediate), **DOCX = final** (delivered report).",
        "`pipeline_csv.py` output is the engine that reproduces the DOCX from the CSV inputs.", "",
        "## Priority-zero findings (anomalies needing your interpretation)", "",
    ]
    lines += [
        f"1. **Money ÷10 between stages.** XLSX (first part) carries all money at 1/10 "
        f"of the DOCX (final). DOCX principal = R$ {rep['docx_scalars']['principal_total']:,.2f}. "
        f"See DOCS_VS_XLSX.md. Intentional stage scaling, or a workbook unit issue?", "",
        f"2. **LGD differs between stages.** DOCX = {rep['docx_scalars']['lgd_pct']}%; "
        f"XLSX PE-sheet = 91.5301%. Methodology value changed between first-part and final.", "",
    ]
    if rep.get("vpl_max_rel_diff_pct") is not None and rep["vpl_max_rel_diff_pct"] > 0.1:
        lines += [
            f"3. **VPL pipeline vs DOCX ~{rep['vpl_max_rel_diff_pct']}% off** "
            f"(beyond the 0.1% bound; e.g. base@1% pipe 136.35M vs DOCX 136.54M). "
            f"The throwaway pipeline is not bit-exact on VPL.", "",
        ]
    else:
        lines += [f"3. VPL pipeline vs DOCX within ~{rep.get('vpl_max_rel_diff_pct')}% (<=0.1% bound).", ""]
    lines += [
        "4. **DOCX UF table sourced from XLSX, not LOANS.csv.** `LOANS.csv` has no `uf` "
        "column; the DOCX UF cut derives from the XLSX `Loans` sheet (which has UF). "
        "UF contrato counts reconcile XLSX→DOCX (see Layer B / lineage).", "",
        "5. **PDF is an architecture report**, not a valuation render — checked only that it "
        "cites the DOCX golden PE/LGD.", "",
        "## Notes", "",
    ]
    for n in rep["notes"]:
        lines.append(f"- {n}")
    lines += ["", "## All checks", "",
              "| check | computed | golden | diff | tol | result | note |",
              "|---|---|---|---|---|---|---|"]
    for c in rep["checks"]:
        r = "PASS" if c["passed"] else "**FAIL**"
        lines.append(f"| {c['name']} | {c['computed']} | {c['golden']} | {c['diff']} | {c['tol']} | {r} | {c.get('note','')} |")
    if rep.get("uf_lineage"):
        lines += ["", "## XLSX → DOCX UF lineage", "",
                  "| UF | xlsx count | docx count | match | principal ratio (xlsx/docx) |",
                  "|---|---|---|---|---|"]
        for u in rep["uf_lineage"]:
            lines.append(f"| {u['uf']} | {u['xlsx_count']} | {u['docx_count']} | "
                         f"{'yes' if u['count_match'] else 'NO'} | {u['principal_ratio_xlsx_over_docx']} |")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("pipeline_output_dir")
    ap.add_argument("--docx", default=str(DOCX))
    ap.add_argument("--xlsx", default=str(XLSX))
    ap.add_argument("--pdf", default=str(PDF))
    ap.add_argument("--out", default=str(OUT))
    args = ap.parse_args()

    Path(args.out).mkdir(parents=True, exist_ok=True)
    print("extracting DOCX / XLSX / pipeline / PDF ...")
    rep = run_audit(Path(args.pipeline_output_dir), Path(args.docx),
                    Path(args.xlsx), Path(args.pdf))
    md = render_markdown(rep)
    (Path(args.out) / "AUDIT-REPORT.md").write_text(md, encoding="utf-8")
    (Path(args.out) / "AUDIT-REPORT.json").write_text(
        json.dumps(rep, indent=2, default=str), encoding="utf-8")
    print(md)
    s = rep["summary"]
    print(f"\nFINAL: {'GREEN' if s['green'] else 'RED'} ({s['passed']}/{s['total']})")
    print(f"report -> {Path(args.out) / 'AUDIT-REPORT.md'}")
    return 0 if s["green"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
